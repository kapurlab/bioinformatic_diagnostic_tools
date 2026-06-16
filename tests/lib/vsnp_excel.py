#!/usr/bin/env python3
"""vsnp_excel.py — reduce a vSNP3 step1 stats workbook to a comparable JSON.

  vsnp_excel.py <run_outdir>

vSNP3 step1 writes <sample>_<date>_stats.xlsx (label/value pairs) and a VCF, but
no JSON headline. This extracts the diagnostic fields into
<outdir>/vsnp_result.json:

  { "reference": "Mycobacterium_AF2122 by Best Reference",   # species/reference call
    "spoligo_octal": "octal-...",                            # TB spoligotype octal
    "spoligo_sb": "SB0265" }                                 # TB spoligotype SB number

Run it with the vSNP3 env Python (it needs openpyxl). Labels are matched wherever
they appear in the sheet; the value is taken from the adjacent cell (right, else
below), so it works whether the stats sheet is row- or column-oriented.
"""
import json
import sys
from pathlib import Path

import openpyxl

LABELS = {
    "reference": "Reference",
    "spoligo_octal": "Spoligotype Octal Code",
    "spoligo_sb": "Spoligotype SB Number",
    "group": "Groups",
}
_LABEL_SET = set(LABELS.values())


def find_value(ws, label):
    """The stats sheet is headers-in-row-1 / values-in-row-2, so the value is the
    cell BELOW the label. Fall back to the cell to the right for a label/value
    column layout. Skip a candidate that is itself another header label."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == label:
                below = ws.cell(row=cell.row + 1, column=cell.column).value
                if below not in (None, "") and str(below).strip() not in _LABEL_SET:
                    return below
                right = ws.cell(row=cell.row, column=cell.column + 1).value
                if right not in (None, "") and str(right).strip() not in _LABEL_SET:
                    return right
                return below
    return None


def main():
    if len(sys.argv) != 2:
        sys.exit("usage: vsnp_excel.py <run_outdir>")
    outdir = Path(sys.argv[1])
    xlsxs = sorted(outdir.glob("*stats.xlsx")) or sorted(outdir.glob("*.xlsx"))
    if not xlsxs:
        sys.exit(f"vsnp_excel.py: no stats .xlsx found under {outdir}")
    ws = openpyxl.load_workbook(xlsxs[0], data_only=True).active
    result = {k: find_value(ws, lab) for k, lab in LABELS.items()}
    result["__xlsx"] = xlsxs[0].name
    (outdir / "vsnp_result.json").write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {outdir/'vsnp_result.json'}: {result}")


if __name__ == "__main__":
    main()
